"""
Google Services for Sheets, Drive, and Excel file operations.
"""

import os
import tempfile
import logging
import json
from typing import List, Dict, Any, Optional, Tuple
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from google.oauth2 import service_account
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.auth.exceptions import DefaultCredentialsError
import pickle
import openpyxl
from openpyxl import Workbook, load_workbook
import csv

logger = logging.getLogger(__name__)

# OAuth2 scopes
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive',
    'https://www.googleapis.com/auth/drive.file'
]

class GoogleService:
    """Service for Google Sheets, Drive, Excel, and CSV operations."""
    
    def __init__(self):
        """Initialize Google service with credentials."""
        self.credentials_file = os.getenv('GOOGLE_CREDENTIALS_FILE', 'credentials.json')
        self.oauth_credentials_file = os.getenv('GOOGLE_OAUTH_CREDENTIALS_FILE', 'oauth_credentials.json')
        self.sheet_id = os.getenv('GOOGLE_SHEET_ID')
        self.drive_folder_id = os.getenv('GOOGLE_DRIVE_FOLDER_ID')
        
        # Initialize Google API clients
        self._init_google_clients()
    
    def _get_oauth_credentials(self):
        """Get OAuth2 credentials with automatic token refresh."""
        creds = None
        
        # Check if we have a token file
        token_file = 'token.pickle'
        if os.path.exists(token_file):
            try:
                with open(token_file, 'rb') as token:
                    creds = pickle.load(token)
                logger.info("Loaded OAuth2 token from file")
            except Exception as e:
                logger.warning(f"Failed to load OAuth2 token: {e}")
        
        # If no valid credentials, get new ones
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try:
                    creds.refresh(Request())
                    logger.info("Refreshed OAuth2 token")
                except Exception as e:
                    logger.warning(f"Failed to refresh OAuth2 token: {e}")
                    creds = None
            
            if not creds:
                # Check if OAuth credentials file exists
                if os.path.exists(self.oauth_credentials_file):
                    try:
                        flow = InstalledAppFlow.from_client_secrets_file(
                            self.oauth_credentials_file, SCOPES)
                        creds = flow.run_local_server(port=0)
                        logger.info("Created new OAuth2 credentials")
                    except Exception as e:
                        logger.error(f"Failed to create OAuth2 credentials: {e}")
                        return None
                else:
                    logger.info("OAuth credentials file not found, falling back to service account")
                    return None
            
            # Save the credentials for next run
            if creds:
                try:
                    with open(token_file, 'wb') as token:
                        pickle.dump(creds, token)
                    logger.info("Saved OAuth2 token to file")
                except Exception as e:
                    logger.warning(f"Failed to save OAuth2 token: {e}")
        
        return creds
    
    def check_credentials(self) -> Dict[str, Any]:
        """
        Check if Google credentials are properly configured.
        
        Returns:
            Dict[str, Any]: Status and details about credentials
        """
        try:
            # Check if credentials file exists
            if not os.path.exists(self.credentials_file):
                return {
                    'status': 'missing_file',
                    'error': f"Google credentials file not found at: {self.credentials_file}",
                    'solution': "Please download your service account JSON file and place it in the credentials folder.",
                    'setup_steps': [
                        "1. Go to Google Cloud Console (https://console.cloud.google.com/)",
                        "2. Create a new project or select existing one",
                        "3. Enable Google Sheets API and Google Drive API",
                        "4. Create a Service Account",
                        "5. Download the JSON key file",
                        "6. Place it in the 'credentials' folder as 'service-account.json'"
                    ]
                }
            
            # Check if credentials file is readable
            try:
                with open(self.credentials_file, 'r') as f:
                    cred_data = json.load(f)
            except json.JSONDecodeError:
                return {
                    'status': 'invalid_json',
                    'error': f"Invalid JSON format in credentials file: {self.credentials_file}",
                    'solution': "Please ensure the credentials file contains valid JSON.",
                    'setup_steps': [
                        "1. Download a fresh copy of your service account JSON file",
                        "2. Ensure the file is not corrupted",
                        "3. Place it in the 'credentials' folder"
                    ]
                }
            except Exception as e:
                return {
                    'status': 'file_error',
                    'error': f"Cannot read credentials file: {str(e)}",
                    'solution': "Please check file permissions and ensure the file is accessible.",
                    'setup_steps': [
                        "1. Check file permissions",
                        "2. Ensure the file is not locked by another process",
                        "3. Try downloading the credentials file again"
                    ]
                }
            
            # Check required fields in credentials
            required_fields = ['type', 'project_id', 'private_key_id', 'private_key', 'client_email']
            missing_fields = [field for field in required_fields if field not in cred_data]
            
            if missing_fields:
                return {
                    'status': 'incomplete_credentials',
                    'error': f"Missing required fields in credentials: {', '.join(missing_fields)}",
                    'solution': "Please ensure your service account JSON file contains all required fields.",
                    'setup_steps': [
                        "1. Download a complete service account JSON file from Google Cloud Console",
                        "2. Ensure the file includes all required fields",
                        "3. Replace the existing credentials file"
                    ]
                }
            
            # Test API connection
            try:
                scopes = [
                    'https://www.googleapis.com/auth/spreadsheets',
                    'https://www.googleapis.com/auth/drive',
                    'https://www.googleapis.com/auth/drive.file'
                ]
                
                credentials = service_account.Credentials.from_service_account_file(
                    self.credentials_file, 
                    scopes=scopes
                )
                
                # Test with a simple API call
                test_service = build('drive', 'v3', credentials=credentials)
                test_service.files().list(pageSize=1).execute()
                
                return {
                    'status': 'valid',
                    'message': "Google credentials are properly configured and working.",
                    'project_id': cred_data.get('project_id', 'Unknown'),
                    'client_email': cred_data.get('client_email', 'Unknown')
                }
                
            except Exception as e:
                error_msg = str(e)
                if 'invalid_grant' in error_msg.lower():
                    return {
                        'status': 'invalid_grant',
                        'error': "Invalid grant - the service account may not have proper permissions",
                        'solution': "Please ensure the service account has the necessary API permissions enabled.",
                        'setup_steps': [
                            "1. Go to Google Cloud Console",
                            "2. Enable Google Sheets API and Google Drive API",
                            "3. Grant the service account appropriate roles",
                            "4. Share your Google Sheets with the service account email"
                        ]
                    }
                elif 'quota' in error_msg.lower():
                    return {
                        'status': 'quota_exceeded',
                        'error': "API quota exceeded - too many requests",
                        'solution': "Please wait before making more requests or upgrade your Google Cloud quota.",
                        'setup_steps': [
                            "1. Wait a few minutes before trying again",
                            "2. Check your Google Cloud Console for quota limits",
                            "3. Consider upgrading your API quota if needed"
                        ]
                    }
                else:
                    return {
                        'status': 'api_error',
                        'error': f"API connection failed: {error_msg}",
                        'solution': "Please check your internet connection and API permissions.",
                        'setup_steps': [
                            "1. Check your internet connection",
                            "2. Verify API services are enabled in Google Cloud Console",
                            "3. Ensure the service account has proper permissions"
                        ]
                    }
                    
        except Exception as e:
            return {
                'status': 'unknown_error',
                'error': f"Unexpected error checking credentials: {str(e)}",
                'solution': "Please check the application logs for more details.",
                'setup_steps': [
                    "1. Check the application logs",
                    "2. Restart the application",
                    "3. Contact support if the issue persists"
                ]
            }
    
    def _init_google_clients(self):
        """Initialize Google API clients."""
        try:
            # Try OAuth2 first
            oauth_creds = self._get_oauth_credentials()
            if oauth_creds:
                # Use OAuth2 credentials
                self.sheets_service = build('sheets', 'v4', credentials=oauth_creds)
                self.drive_service = build('drive', 'v3', credentials=oauth_creds)
                logger.info("Google API clients initialized with OAuth2")
                return
            
            # Fall back to service account
            logger.info("OAuth2 not available, using service account")
            
            # Check credentials first
            cred_status = self.check_credentials()
            if cred_status['status'] != 'valid':
                raise Exception(f"Google credentials issue: {cred_status['error']}")
            
            # Define scopes
            scopes = [
                'https://www.googleapis.com/auth/spreadsheets',
                'https://www.googleapis.com/auth/drive',
                'https://www.googleapis.com/auth/drive.file'
            ]
            
            # Load credentials
            credentials = service_account.Credentials.from_service_account_file(
                self.credentials_file, 
                scopes=scopes
            )
            
            # Build API clients
            self.sheets_service = build('sheets', 'v4', credentials=credentials)
            self.drive_service = build('drive', 'v3', credentials=credentials)
            
            logger.info("Google API clients initialized with service account")
            
        except Exception as e:
            logger.error(f"Failed to initialize Google API clients: {e}")
            raise
    
    def read_sheet_data(self, sheet_id: str = None, range_name: str = "A1:Z10") -> List[List[str]]:
        """
        Read data from Google Sheets.
        
        Args:
            sheet_id (str, optional): Sheet ID, uses default if not provided
            range_name (str): Range to read (e.g., "A1:B10")
            
        Returns:
            List[List[str]]: Sheet data
        """
        try:
            sheet_id = sheet_id or self.sheet_id
            if not sheet_id:
                raise ValueError("Sheet ID is required")
            
            # Validate range format
            self._validate_range(range_name)
            
            # Get data from sheet
            result = self.sheets_service.spreadsheets().values().get(
                spreadsheetId=sheet_id, 
                range=range_name
            ).execute()
            
            values = result.get('values', [])
            logger.info(f"Read {len(values)} rows from sheet {sheet_id}")
            
            return values
            
        except Exception as e:
            logger.error(f"Error reading sheet data: {e}")
            raise
    
    def write_sheet_data(self, sheet_id: str = None, range_name: str = None, 
                        values: List[List[str]] = None, value: str = None) -> bool:
        """
        Write data to Google Sheets.
        
        Args:
            sheet_id (str, optional): Sheet ID, uses default if not provided
            range_name (str): Range to write to (e.g., "A1")
            values (List[List[str]], optional): 2D array of values
            value (str, optional): Single value to write
            
        Returns:
            bool: True if successful
        """
        try:
            sheet_id = sheet_id or self.sheet_id
            if not sheet_id:
                raise ValueError("Sheet ID is required")
            
            if not range_name:
                raise ValueError("Range is required")
            
            # Validate range format
            self._validate_range(range_name)
            
            # Prepare data
            if value is not None:
                data = [[value]]
            elif values is not None:
                data = values
            else:
                raise ValueError("Either value or values must be provided")
            
            # Write data
            body = {'values': data}
            self.sheets_service.spreadsheets().values().update(
                spreadsheetId=sheet_id,
                range=range_name,
                valueInputOption='RAW',
                body=body
            ).execute()
            
            logger.info(f"Wrote data to sheet {sheet_id} at range {range_name}")
            return True
            
        except Exception as e:
            logger.error(f"Error writing sheet data: {e}")
            raise
    
    def create_new_sheet(self, title: str, data: List[List[str]] = None) -> Dict[str, Any]:
        """
        Create a new Google Sheet.
        
        Args:
            title (str): Sheet title
            data (List[List[str]], optional): Initial data
            
        Returns:
            Dict[str, Any]: Sheet information
        """
        try:
            # Create spreadsheet
            spreadsheet = {
                'properties': {
                    'title': title
                },
                'sheets': [
                    {
                        'properties': {
                            'title': 'Sheet1'
                        }
                    }
                ]
            }
            
            result = self.sheets_service.spreadsheets().create(body=spreadsheet).execute()
            sheet_id = result['spreadsheetId']
            
            # Add initial data if provided
            if data:
                self.write_sheet_data(sheet_id, "A1", data)
            
            # Set permissions if folder is specified
            if self.drive_folder_id:
                self._move_to_folder(sheet_id, self.drive_folder_id)
            
            logger.info(f"Created new sheet: {title} (ID: {sheet_id})")
            
            return {
                'id': sheet_id,
                'title': title,
                'url': f"https://docs.google.com/spreadsheets/d/{sheet_id}"
            }
            
        except Exception as e:
            logger.error(f"Error creating new sheet: {e}")
            raise
    
    def read_csv_from_drive(self, filename: str) -> List[List[str]]:
        """
        Read CSV file from Google Drive.
        
        Args:
            filename (str): CSV filename
            
        Returns:
            List[List[str]]: CSV data
        """
        try:
            # Find file in Drive
            file_id = self._get_drive_file_id(filename, 'text/csv')
            
            # Download file
            local_path = self._download_drive_file(file_id)
            
            # Read CSV data
            data = []
            with open(local_path, 'r', newline='', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                data = list(reader)
            
            # Clean up
            os.unlink(local_path)
            
            logger.info(f"Read {len(data)} rows from CSV file: {filename}")
            return data
            
        except Exception as e:
            logger.error(f"Error reading CSV from Drive: {e}")
            raise
    
    def write_csv_to_drive(self, filename: str, data: List[List[str]], 
                          create_new: bool = False) -> bool:
        """
        Write CSV data to Google Drive.
        
        Args:
            filename (str): CSV filename
            data (List[List[str]]): CSV data
            create_new (bool): Create new file if doesn't exist
            
        Returns:
            bool: True if successful
        """
        try:
            # Create temporary file
            with tempfile.NamedTemporaryFile(mode='w', newline='', suffix='.csv', delete=False) as temp_file:
                writer = csv.writer(temp_file)
                writer.writerows(data)
                temp_path = temp_file.name
            
            try:
                if create_new:
                    # Create new file in Drive
                    file_metadata = {
                        'name': filename,
                        'parents': [self.drive_folder_id] if self.drive_folder_id else []
                    }
                    
                    media = MediaFileUpload(temp_path, mimetype='text/csv', resumable=True)
                    file = self.drive_service.files().create(
                        body=file_metadata,
                        media_body=media,
                        fields='id'
                    ).execute()
                    
                    logger.info(f"Created new CSV file: {filename} (ID: {file['id']})")
                else:
                    # Update existing file
                    file_id = self._get_drive_file_id(filename, 'text/csv')
                    
                    media = MediaFileUpload(temp_path, mimetype='text/csv', resumable=True)
                    self.drive_service.files().update(
                        fileId=file_id,
                        media_body=media
                    ).execute()
                    
                    logger.info(f"Updated CSV file: {filename}")
                
                return True
                
            finally:
                # Clean up temporary file
                os.unlink(temp_path)
                
        except Exception as e:
            logger.error(f"Error writing CSV to Drive: {e}")
            raise
    
    def read_excel_from_drive(self, filename: str, sheet_name: str = None) -> List[List[str]]:
        """
        Read Excel file from Google Drive.
        
        Args:
            filename (str): Excel filename
            sheet_name (str, optional): Sheet name, uses first sheet if not provided
            
        Returns:
            List[List[str]]: Excel data
        """
        try:
            # Find file in Drive
            file_id = self._get_drive_file_id(filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            
            # Download file
            local_path = self._download_drive_file(file_id)
            
            # Read Excel data
            workbook = load_workbook(local_path, read_only=True)
            sheet = workbook[sheet_name] if sheet_name else workbook.active
            
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append([str(cell) if cell is not None else '' for cell in row])
            
            workbook.close()
            
            # Clean up
            os.unlink(local_path)
            
            logger.info(f"Read {len(data)} rows from Excel file: {filename}")
            return data
            
        except Exception as e:
            logger.error(f"Error reading Excel from Drive: {e}")
            raise
    
    def write_excel_to_drive(self, filename: str, data: List[List[str]], 
                            sheet_name: str = "Sheet1", create_new: bool = False) -> bool:
        """
        Write Excel data to Google Drive.
        
        Args:
            filename (str): Excel filename
            data (List[List[str]]): Excel data
            sheet_name (str): Sheet name
            create_new (bool): Create new file if doesn't exist
            
        Returns:
            bool: True if successful
        """
        try:
            # Create temporary file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_path = temp_file.name
            
            try:
                # Create workbook
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = sheet_name
                
                # Write data
                for row_idx, row_data in enumerate(data, 1):
                    for col_idx, cell_value in enumerate(row_data, 1):
                        sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Save workbook
                workbook.save(temp_path)
                workbook.close()
                
                if create_new:
                    # Create new file in Drive
                    file_metadata = {
                        'name': filename,
                        'parents': [self.drive_folder_id] if self.drive_folder_id else []
                    }
                    
                    media = MediaFileUpload(temp_path, 
                                          mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                          resumable=True)
                    file = self.drive_service.files().create(
                        body=file_metadata,
                        media_body=media,
                        fields='id'
                    ).execute()
                    
                    logger.info(f"Created new Excel file: {filename} (ID: {file['id']})")
                else:
                    # Update existing file
                    file_id = self._get_drive_file_id(filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    
                    media = MediaFileUpload(temp_path,
                                          mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                          resumable=True)
                    self.drive_service.files().update(
                        fileId=file_id,
                        media_body=media
                    ).execute()
                    
                    logger.info(f"Updated Excel file: {filename}")
                
                return True
                
            finally:
                # Clean up temporary file
                os.unlink(temp_path)
                
        except Exception as e:
            logger.error(f"Error writing Excel to Drive: {e}")
            raise
    
    def _get_drive_file_id(self, filename: str, mime_type: str) -> str:
        """Get file ID from Google Drive by filename and MIME type."""
        try:
            query = f"name='{filename}' and mimeType='{mime_type}'"
            if self.drive_folder_id:
                query += f" and '{self.drive_folder_id}' in parents"
            
            results = self.drive_service.files().list(
                q=query,
                spaces='drive',
                fields="files(id, name)"
            ).execute()
            
            files = results.get('files', [])
            if not files:
                raise FileNotFoundError(f"File '{filename}' not found in Drive")
            
            return files[0]['id']
            
        except Exception as e:
            logger.error(f"Error getting Drive file ID: {e}")
            raise
    
    def _download_drive_file(self, file_id: str) -> str:
        """Download file from Google Drive to temporary location."""
        try:
            request = self.drive_service.files().get_media(fileId=file_id)
            fh = tempfile.NamedTemporaryFile(delete=False)
            
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            
            while not done:
                status, done = downloader.next_chunk()
            
            fh.close()
            return fh.name
            
        except Exception as e:
            logger.error(f"Error downloading Drive file: {e}")
            raise
    
    def _move_to_folder(self, file_id: str, folder_id: str):
        """Move file to specified folder."""
        try:
            file = self.drive_service.files().get(fileId=file_id, fields='parents').execute()
            previous_parents = ",".join(file.get('parents', []))
            
            self.drive_service.files().update(
                fileId=file_id,
                addParents=folder_id,
                removeParents=previous_parents,
                fields='id, parents'
            ).execute()
            
        except Exception as e:
            logger.error(f"Error moving file to folder: {e}")
            raise
    
    def _validate_range(self, range_name: str):
        """Validate Google Sheets range format."""
        import re
        if not re.match(r'^[A-Z]+[0-9]+(:[A-Z]+[0-9]+)?$', range_name):
            raise ValueError(f"Invalid range format: {range_name}")
    
    def get_file_info(self, filename: str) -> Dict[str, Any]:
        """
        Get file information from Google Drive.
        
        Args:
            filename (str): Filename
            
        Returns:
            Dict[str, Any]: File information
        """
        try:
            # Try different MIME types
            mime_types = [
                'text/csv',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'application/vnd.ms-excel'
            ]
            
            for mime_type in mime_types:
                try:
                    file_id = self._get_drive_file_id(filename, mime_type)
                    
                    file_info = self.drive_service.files().get(
                        fileId=file_id,
                        fields='id,name,mimeType,createdTime,modifiedTime,size'
                    ).execute()
                    
                    return {
                        'id': file_info['id'],
                        'name': file_info['name'],
                        'mime_type': file_info['mimeType'],
                        'created_time': file_info['createdTime'],
                        'modified_time': file_info['modifiedTime'],
                        'size': file_info.get('size', 0)
                    }
                    
                except FileNotFoundError:
                    continue
            
            raise FileNotFoundError(f"File '{filename}' not found in Drive")
            
        except Exception as e:
            logger.error(f"Error getting file info: {e}")
            raise
    
    def list_available_sheets(self) -> List[Dict[str, Any]]:
        """
        List all available Google Sheets in the specified folder.
        
        Returns:
            List[Dict[str, Any]]: List of sheet information
        """
        try:
            # Build query to search for Google Sheets in the specific folder
            if self.drive_folder_id:
                query = f"mimeType='application/vnd.google-apps.spreadsheet' and trashed=false and '{self.drive_folder_id}' in parents"
                logger.info(f"Searching for sheets in folder: {self.drive_folder_id}")
            else:
                query = "mimeType='application/vnd.google-apps.spreadsheet' and trashed=false"
                logger.warning("No GOOGLE_DRIVE_FOLDER_ID specified, searching all sheets")
            
            results = self.drive_service.files().list(
                q=query,
                fields="files(id,name,createdTime,modifiedTime,webViewLink)",
                orderBy="modifiedTime desc"
            ).execute()
            
            files = results.get('files', [])
            logger.info(f"Found {len(files)} Google Sheets in folder")
            
            sheets = []
            for file in files:
                sheets.append({
                    'id': file['id'],
                    'name': file['name'],
                    'url': file['webViewLink'],
                    'created': file['createdTime'],
                    'modified': file['modifiedTime']
                })
            
            return sheets
            
        except Exception as e:
            logger.error(f"Error listing available sheets: {e}")
            return []
    
    def list_available_excel_files(self) -> List[Dict[str, Any]]:
        """
        List all available Excel files in the specified folder.
        
        Returns:
            List[Dict[str, Any]]: List of Excel file information
        """
        try:
            # Build query to search for Excel files in the specific folder
            if self.drive_folder_id:
                query = f"(mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or mimeType='application/vnd.ms-excel') and trashed=false and '{self.drive_folder_id}' in parents"
                logger.info(f"Searching for Excel files in folder: {self.drive_folder_id}")
            else:
                query = "(mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or mimeType='application/vnd.ms-excel') and trashed=false"
                logger.warning("No GOOGLE_DRIVE_FOLDER_ID specified, searching all Excel files")
            
            results = self.drive_service.files().list(
                q=query,
                fields="files(id,name,createdTime,modifiedTime,webViewLink)",
                orderBy="modifiedTime desc"
            ).execute()
            
            files = results.get('files', [])
            logger.info(f"Found {len(files)} Excel files in folder")
            
            excel_files = []
            for file in files:
                excel_files.append({
                    'id': file['id'],
                    'name': file['name'],
                    'url': file['webViewLink'],
                    'created': file['createdTime'],
                    'modified': file['modifiedTime']
                })
            
            return excel_files
            
        except Exception as e:
            logger.error(f"Error listing available Excel files: {e}")
            return []
    
    def list_available_csv_files(self) -> List[Dict[str, Any]]:
        """
        List all available CSV files in the specified folder.
        
        Returns:
            List[Dict[str, Any]]: List of CSV file information
        """
        try:
            # Build query to search for CSV files in the specific folder
            if self.drive_folder_id:
                query = f"mimeType='text/csv' and trashed=false and '{self.drive_folder_id}' in parents"
                logger.info(f"Searching for CSV files in folder: {self.drive_folder_id}")
            else:
                query = "mimeType='text/csv' and trashed=false"
                logger.warning("No GOOGLE_DRIVE_FOLDER_ID specified, searching all CSV files")
            
            results = self.drive_service.files().list(
                q=query,
                fields="files(id,name,createdTime,modifiedTime,webViewLink)",
                orderBy="modifiedTime desc"
            ).execute()
            
            files = results.get('files', [])
            logger.info(f"Found {len(files)} CSV files in folder")
            
            csv_files = []
            for file in files:
                csv_files.append({
                    'id': file['id'],
                    'name': file['name'],
                    'url': file['webViewLink'],
                    'created': file['createdTime'],
                    'modified': file['modifiedTime']
                })
            
            return csv_files
            
        except Exception as e:
            logger.error(f"Error listing available CSV files: {e}")
            return []
    
 